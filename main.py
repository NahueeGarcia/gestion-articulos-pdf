import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from reportlab.pdfgen import canvas
import os
import pandas as pd
from reportlab.lib.pagesizes import letter
from datetime import datetime
from openpyxl import load_workbook

def crear_pdf_articulo(nombre_archivo, datos_por_articulo, carpeta_destino="pdfs"):
    # Crear la carpeta si no existe
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    # Ruta completa del archivo
    ruta_completa = os.path.join(carpeta_destino, f"{nombre_archivo}.pdf")
    
    # Crear el PDF
    c = canvas.Canvas(ruta_completa, pagesize=letter)
    width, height = letter
    
    # Configuración inicial
    y_position = height - 50
    row_height = 15  # mismo que pasillo
    font_header = 9
    font_data = 8

    # Columnas (idénticas a pasillo)
    col_localizador = 50
    col_articulo = 120
    col_descripcion = 190
    col_en_mano = 390
    col_lpn = 460
    columns = [col_localizador, col_articulo, col_descripcion, col_en_mano, col_lpn]
    
    headers = ["Localizador", "Artículo", "Descripción", "En Mano", "LPN"]

    if not datos_por_articulo:
        c.setFont("Helvetica", font_data)
        c.drawString(50, y_position, "No se encontraron datos para este artículo.")
        c.save()
        return ruta_completa

    # Función para dibujar headers
    def dibujar_headers(y_pos):
        c.setFont("Helvetica-Bold", font_header)
        for i, (header, col_x) in enumerate(zip(headers, columns)):
            c.setFillColorRGB(0.9, 0.9, 0.9)
            if i == 0:
                c.rect(col_x, y_pos - row_height + 5, columns[1] - col_x, row_height, fill=1, stroke=1)
            elif i == len(headers) - 1:
                c.rect(col_x, y_pos - row_height + 5, width - 50 - col_x, row_height, fill=1, stroke=1)
            else:
                c.rect(col_x, y_pos - row_height + 5, columns[i+1] - col_x, row_height, fill=1, stroke=1)
            c.setFillColorRGB(0, 0, 0)
            c.drawString(col_x + 2, y_pos - 8, header)  # alineado igual que pasillo

    # Dibujar headers iniciales
    dibujar_headers(y_position)
    y_position -= row_height

    # Filas
    c.setFont("Helvetica", font_data)
    for i, (articulo, descripcion, en_mano, localizador, lpn) in enumerate(datos_por_articulo):
        if y_position < 60:  # salto de página
            c.showPage()
            y_position = height - 50
            dibujar_headers(y_position)
            y_position -= row_height
            c.setFont("Helvetica", font_data)

        # Fondo alternado
        c.setFillColorRGB(0.98, 0.98, 0.98) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
        for j, col_x in enumerate(columns):
            if j == 0:
                c.rect(col_x, y_position - row_height + 5, columns[1] - col_x, row_height, fill=1, stroke=1)
            elif j == len(columns) - 1:
                c.rect(col_x, y_position - row_height + 5, width - 50 - col_x, row_height, fill=1, stroke=1)
            else:
                c.rect(col_x, y_position - row_height + 5, columns[j+1] - col_x, row_height, fill=1, stroke=1)

        c.setFillColorRGB(0, 0, 0)

        # Stock
        try:
            stock_texto = str(en_mano) if en_mano != '' else '0'
        except (ValueError, TypeError):
            stock_texto = '0'
        
        # LPN
        if pd.isna(lpn) or str(lpn).lower() == 'nan':
            lpn_texto = "-"
        else:
            lpn_texto = str(lpn)
        
        # Descripción más corta para encajar
        desc_truncada = str(descripcion)[:35] + "..." if len(str(descripcion)) > 35 else str(descripcion)

        # Escribir datos (mismo offset que pasillo → -8)
        c.drawString(col_localizador + 2, y_position - 8, str(localizador))
        c.drawString(col_articulo + 2, y_position - 8, str(articulo))
        c.drawString(col_descripcion + 2, y_position - 8, desc_truncada)
        c.drawString(col_en_mano + 2, y_position - 8, stock_texto)
        c.drawString(col_lpn + 2, y_position - 8, lpn_texto)

        y_position -= row_height

    c.save()
    return ruta_completa


def mostrar_mensaje_exito(mensaje, label_mensaje):
    label_mensaje.configure(text=mensaje, font=('Arial', 10, 'bold'))
    
    # Limpiar después de 3 segundos
    label_mensaje.after(3000, lambda: label_mensaje.configure(text="", font=('Arial', 10)))

def mostrar_mensaje_descripcion(descripcion, label_mensaje):
    label_mensaje.configure(text=descripcion, font=('Arial', 10, 'bold'))
    label_mensaje.pack(pady=10)  # usar pack en vez de grid/place

    # limpiar después de 30 segundos
    label_mensaje.after(30000, lambda: label_mensaje.configure(text="", font=('Arial', 10)))

def obtener_datos_por_articulo(articulo, archivo_excel):
    resultados = []
    
    try:
        # Si es un StringVar de tkinter, extraer el valor
        if hasattr(archivo_excel, 'get'):
            ruta_archivo = archivo_excel.get()
        else:
            ruta_archivo = archivo_excel
        
        # Verificar que la ruta no esté vacía
        if not ruta_archivo:
            print("Error: No se ha seleccionado ningún archivo")
            return []
        
        # CORRECCIÓN: Usar ruta_archivo en lugar de archivo_excel
        todas_las_hojas = pd.read_excel(ruta_archivo, sheet_name=None)
        
        # Iterar sobre cada hoja
        for nombre_hoja, df in todas_las_hojas.items():
            # Verificar que la hoja tenga las columnas necesarias
            if 'Artículo' in df.columns:
                # Convertir la columna Artículo a string para evitar problemas con floats
                df['Artículo'] = df['Artículo'].astype(str)
                # Limpiar decimales innecesarios (ej: "426367.0" -> "426367")
                df['Artículo'] = df['Artículo'].str.replace(r'\.0$', '', regex=True)
                
                # Convertir el parámetro articulo a string también
                articulo_str = str(articulo)
                
                # Buscar filas donde la columna 'Artículo' coincida con el parámetro
                coincidencias = df[df['Artículo'] == articulo_str]
                
                # Para cada coincidencia, extraer los datos en el orden solicitado
                for _, fila in coincidencias.iterrows():
                    resultado = (
                        fila.get('Artículo', ''),           # Columna D
                        fila.get('Desc Artículo', ''),      # Columna E  
                        fila.get('En Mano', ''),            # Columna H
                        fila.get('Localizador', ''),        # Columna C
                        fila.get('LPN', '')                 # Columna O
                    )
                    resultados.append(resultado)
        
        return resultados
        
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {ruta_archivo}")
        return []
    except Exception as e:
        print(f"Error al procesar el archivo: {str(e)}")
        return []

def crear_pdf_pasillo(pasillo, datos_por_pasillo, carpeta_destino="pdfs"):
    
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    ruta_completa = os.path.join(carpeta_destino, f"pasillo{pasillo}.pdf")
    c = canvas.Canvas(ruta_completa, pagesize=letter)
    width, height = letter
    
    # Configuración inicial
    y_position = height - 50
    row_height = 15  # compacto pero legible
    font_header = 9
    font_data = 8

    if not datos_por_pasillo:
        c.setFont("Helvetica", font_data)
        c.drawString(50, y_position, f"No se encontraron datos para el pasillo {pasillo}.")
        c.save()
        return ruta_completa
    
    # Columnas
    col_localizador = 50
    col_articulo = 120
    col_descripcion = 190
    col_en_mano = 390
    col_lpn = 460
    columns = [col_localizador, col_articulo, col_descripcion, col_en_mano, col_lpn]
    
    headers = ["Localizador", "Artículo", "Descripción", "En Mano", "LPN"]
    
    def dibujar_headers(y_pos):
        c.setFont("Helvetica-Bold", font_header)
        for i, (header, col_x) in enumerate(zip(headers, columns)):
            c.setFillColorRGB(0.9, 0.9, 0.9)
            if i == 0:
                c.rect(col_x, y_pos - row_height + 5, columns[1] - col_x, row_height, fill=1, stroke=1)
            elif i == len(headers) - 1:
                c.rect(col_x, y_pos - row_height + 5, width - 50 - col_x, row_height, fill=1, stroke=1)
            else:
                c.rect(col_x, y_pos - row_height + 5, columns[i+1] - col_x, row_height, fill=1, stroke=1)
            c.setFillColorRGB(0, 0, 0)
            # 🔹 Subido un poco para que no quede pisado
            c.drawString(col_x + 2, y_pos - 8, header)
    
    # Dibujar encabezados iniciales
    dibujar_headers(y_position)
    y_position -= row_height
    
    # Filas
    c.setFont("Helvetica", font_data)
    for i, (localizador, articulo, desc, en_mano, lpn) in enumerate(datos_por_pasillo):
        if y_position < 60:  # salto de página
            c.showPage()
            y_position = height - 50
            dibujar_headers(y_position)
            y_position -= row_height
            c.setFont("Helvetica", font_data)
        
        # Fondo alternado
        c.setFillColorRGB(0.98, 0.98, 0.98) if i % 2 == 0 else c.setFillColorRGB(1, 1, 1)
        for j, col_x in enumerate(columns):
            if j == 0:
                c.rect(col_x, y_position - row_height + 5, columns[1] - col_x, row_height, fill=1, stroke=1)
            elif j == len(columns) - 1:
                c.rect(col_x, y_position - row_height + 5, width - 50 - col_x, row_height, fill=1, stroke=1)
            else:
                c.rect(col_x, y_position - row_height + 5, columns[j+1] - col_x, row_height, fill=1, stroke=1)
        
        # Texto procesado
        articulo_texto = "" if pd.isna(articulo) else str(articulo).replace('.0', '') if str(articulo).endswith('.0') else str(articulo)
        stock_texto = "0" if pd.isna(en_mano) or en_mano == '' else str(en_mano)
        lpn_texto = "-" if pd.isna(lpn) or str(lpn).lower() == 'nan' else str(lpn)
        desc_truncada = str(desc)[:35] + "..." if len(str(desc)) > 35 else str(desc)
        
        c.setFillColorRGB(0, 0, 0)
        # 🔹 Subido un poco (antes -10, ahora -8)
        c.drawString(col_localizador + 2, y_position - 8, str(localizador))
        c.drawString(col_articulo + 2, y_position - 8, articulo_texto)
        c.drawString(col_descripcion + 2, y_position - 8, desc_truncada)
        c.drawString(col_en_mano + 2, y_position - 8, stock_texto)
        c.drawString(col_lpn + 2, y_position - 8, lpn_texto)
        
        y_position -= row_height
    
    c.save()
    return ruta_completa


def obtener_datos_por_pasillo(pasillo, archivo_excel):
    resultados = []
    
    try:
        # Si es un StringVar de tkinter, extraer el valor
        if hasattr(archivo_excel, 'get'):
            ruta_archivo = archivo_excel.get()
        else:
            ruta_archivo = archivo_excel
        
        # Verificar que la ruta no esté vacía
        if not ruta_archivo:
            print("Error: No se ha seleccionado ningún archivo")
            return []
        
        # Formatear el pasillo con ceros a la izquierda (ej: "2" -> "P02")
        pasillo_formateado = f"P{int(pasillo):02d}"
        
        # Leer todas las hojas del archivo Excel
        todas_las_hojas = pd.read_excel(ruta_archivo, sheet_name=None)
        
        # Lista temporal para poder ordenar después
        datos_temporales = []
        
        # Iterar sobre cada hoja
        for nombre_hoja, df in todas_las_hojas.items():
            # Verificar que la hoja tenga las columnas necesarias
            if 'Localizador' in df.columns:
                # Convertir localizador a string para poder hacer comparaciones
                df['Localizador'] = df['Localizador'].astype(str)
                
                # Filtrar filas que pertenezcan al pasillo
                filas_pasillo = df[df['Localizador'].str.startswith(pasillo_formateado, na=False)]
                
                # Para cada fila del pasillo, extraer los datos
                for _, fila in filas_pasillo.iterrows():
                    localizador = fila.get('Localizador', '')
                    
                    # Extraer altura y posición para ordenar
                    try:
                        # Formato: P02.041.3.1 -> dividir en partes
                        partes = localizador.split('.')
                        if len(partes) >= 4:
                            posicion = int(partes[1])  # 041
                            altura = int(partes[2])    # 3
                        else:
                            posicion = 999  # Para casos raros, ponerlos al final
                            altura = 999
                    except (ValueError, IndexError):
                        posicion = 999
                        altura = 999
                    
                    resultado = (
                        localizador,                        # Columna C
                        fila.get('Artículo', ''),           # Columna D
                        fila.get('Desc Artículo', ''),      # Columna E  
                        fila.get('En Mano', ''),            # Columna H
                        fila.get('LPN', ''),                # Columna O
                        altura,                             # Para ordenar
                        posicion                            # Para ordenar
                    )
                    datos_temporales.append(resultado)
        
        # Ordenar por altura y después por posición
        datos_temporales.sort(key=lambda x: (x[5], x[6]))  # altura, posición
        
        # Quitar las columnas de ordenamiento y retornar solo los datos necesarios
        resultados = [(loc, art, desc, stock, lpn) for loc, art, desc, stock, lpn, _, _ in datos_temporales]
        
        return resultados
    
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {ruta_archivo}")
        return []
    except Exception as e:
        print(f"Error al procesar el archivo: {str(e)}")
        return []
    
def obtener_descripcion(articulo, archivo_excel):
    try:
        # Si es un StringVar de tkinter, extraer el valor
        if hasattr(archivo_excel, 'get'):
            ruta_archivo = archivo_excel.get()
        else:
            ruta_archivo = archivo_excel

        # Verificar que la ruta no esté vacía
        if not ruta_archivo:
            print("Error: No se ha seleccionado ningún archivo")
            return ""

        # Leer todas las hojas del Excel
        todas_las_hojas = pd.read_excel(ruta_archivo, sheet_name=None)

        # Convertir a string el artículo que buscamos
        articulo_str = str(articulo)

        # Buscar en cada hoja
        for nombre_hoja, df in todas_las_hojas.items():
            if 'Artículo' in df.columns and 'Desc Artículo' in df.columns:
                # Convertir la columna Artículo a string
                df['Artículo'] = df['Artículo'].astype(str)
                df['Artículo'] = df['Artículo'].str.replace(r'\.0$', '', regex=True)

                # Filtrar coincidencias
                coincidencias = df[df['Artículo'] == articulo_str]

                if not coincidencias.empty:
                    # Devolver la primera descripción encontrada
                    return str(coincidencias.iloc[0]['Desc Artículo'])

        # Si no se encontró nada
        return ""

    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {ruta_archivo}")
        return ""
    except Exception as e:
        print(f"Error al procesar el archivo: {str(e)}")
        return ""

def buscar_por_articulo(entry_articulo, archivo_excel, label_mensaje):
    """Crea PDF para artículo específico"""
    articulo = entry_articulo.get().strip()
    
    if not articulo:
        messagebox.showwarning("Advertencia", "Por favor ingresa el nombre del artículo")
        return
    
    try:
        nombre_archivo = f"Articulo{articulo}"
        datos_por_articulo = obtener_datos_por_articulo(articulo, archivo_excel)
        ruta_pdf = crear_pdf_articulo(nombre_archivo, datos_por_articulo)
        mensaje = f"✓ PDF creado exitosamente: {nombre_archivo}.pdf"
        mostrar_mensaje_exito(mensaje, label_mensaje)
    except Exception as e:
        messagebox.showerror("Error", f"Error al crear el PDF: {str(e)}")

def buscar_descripcion_articulo(entry_articulo, archivo_excel, label_mensaje):
    """Crea PDF para artículo específico"""
    articulo = entry_articulo.get().strip()
    
    if not articulo:
        messagebox.showwarning("Advertencia", "Por favor ingresa el nombre del artículo")
        return
    
    try:
        descripcion = obtener_descripcion(articulo, archivo_excel)
        mostrar_mensaje_descripcion(descripcion, label_mensaje)
    except Exception as e:
        messagebox.showerror("Error", f"Error al buscar artículo: {str(e)}")

def buscar_por_pasillo(entry_pasillo, archivo_excel, label_mensaje):
    """Crea PDF para pasillo específico"""
    pasillo = entry_pasillo.get().strip()
    
    if not pasillo:
        messagebox.showwarning("Advertencia", "Por favor ingresa el número de pasillo")
        return
    
    try:
        nombre_archivo = f"Pasillo{pasillo}"
        datos_por_pasillo = obtener_datos_por_pasillo(pasillo, archivo_excel)
        ruta_pdf = crear_pdf_pasillo(pasillo, datos_por_pasillo)
        mensaje = f"✓ PDF creado exitosamente: {nombre_archivo}.pdf"
        mostrar_mensaje_exito(mensaje, label_mensaje)
    except Exception as e:
        messagebox.showerror("Error", f"Error al crear el PDF: {str(e)}")

# --- Función para buscar archivo Excel
def seleccionar_archivo(archivo_excel, lbl_archivo):
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    if archivo:
        archivo_excel.set(archivo)
        lbl_archivo.config(text=f"Archivo: {archivo}")

def obtener_datos_por_localizador(localizador, archivo_excel):
    resultados = []
    
    try:
        # Si es un StringVar de tkinter, extraer el valor
        if hasattr(archivo_excel, 'get'):
            ruta_archivo = archivo_excel.get()
        else:
            ruta_archivo = archivo_excel

        # Verificar que la ruta no esté vacía
        if not ruta_archivo:
            print("Error: No se ha seleccionado ningún archivo")
            return []

        # Leer todas las hojas del Excel
        todas_las_hojas = pd.read_excel(ruta_archivo, sheet_name=None)

        # Convertir a string el localizador buscado
        localizador_str = str(localizador).strip()

        # Buscar en cada hoja
        for nombre_hoja, df in todas_las_hojas.items():
            # Verificar que existan las columnas necesarias
            columnas_necesarias = ['Localizador', 'Artículo', 'Desc Artículo', 'En Mano', 'LPN']
            if all(col in df.columns for col in columnas_necesarias):
                # Convertir la columna Localizador a string para comparar sin problemas
                df['Localizador'] = df['Localizador'].astype(str).str.strip()

                # Filtrar coincidencias
                coincidencias = df[df['Localizador'] == localizador_str]

                # Extraer resultados en el orden solicitado
                for _, fila in coincidencias.iterrows():
                    resultado = (
                        fila.get('Localizador', ''),
                        fila.get('Artículo', ''),
                        fila.get('Desc Artículo', ''),
                        fila.get('En Mano', ''),
                        fila.get('LPN', '')
                    )
                    resultados.append(resultado)

        return resultados

    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {ruta_archivo}")
        return []
    except Exception as e:
        print(f"Error al procesar el archivo: {str(e)}")
        return []

# --- Función para mostrar el formulario correcto
def mostrar_formulario(form_container, modo_var, form_buscar_por_articulo, form_buscar_por_pasillo, form_diferencias):
    for child in form_container.winfo_children():
        child.pack_forget()
    if modo_var.get() == "Buscar por artículo":
        form_buscar_por_articulo.pack(fill="x")
    elif modo_var.get() == "Buscar por pasillo":
        form_buscar_por_pasillo.pack(fill="x")
    elif modo_var.get() == "Diferencias":
        form_diferencias.pack(fill="x")

def agregar_diferencia(datos, estado, archivo_excel):
    try:
        # Si es un StringVar de tkinter, extraer el valor
        if hasattr(archivo_excel, 'get'):
            ruta_archivo = archivo_excel.get()
        else:
            ruta_archivo = archivo_excel

        if not ruta_archivo:
            print("Error: No se ha seleccionado ningún archivo")
            return

        # Intentar abrir el Excel existente
        try:
            book = load_workbook(ruta_archivo)
        except FileNotFoundError:
            print(f"Error: No se encontró el archivo {ruta_archivo}")
            return

        # Nombre de la hoja donde vamos a guardar las diferencias
        hoja_diferencias = "Diferencias"

        # Si la hoja no existe, crearla con encabezados
        if hoja_diferencias not in book.sheetnames:
            ws = book.create_sheet(hoja_diferencias)
            ws.append(["Localizador", "Artículo", "Desc Artículo", "En Mano", "LPN", "Estado"])
        else:
            ws = book[hoja_diferencias]

        # Agregar cada fila de datos
        for fila in datos:
            # fila es: (Localizador, Artículo, Desc Artículo, En Mano, LPN)
            nueva_fila = list(fila) + [estado]  # agregamos la columna "Estado"
            ws.append(nueva_fila)

        # Guardar cambios
        book.save(ruta_archivo)

    except Exception as e:
        print(f"Error al agregar diferencias: {str(e)}")

def agregar_diferencias(entry_diferencias, estado_var, archivo_excel, label_mensaje):
    """Crea PDF para artículo específico"""
    localizador = entry_diferencias.get().strip()
    estado = estado_var.get().strip()
    
    if not localizador or not estado:
        messagebox.showwarning("Advertencia", "Por favor ingresa los campos faltantes")
        return
    
    try:
        datos = obtener_datos_por_localizador(localizador, archivo_excel)
        agregar_diferencia(datos, estado, archivo_excel)
        mensaje = "Diferencia agregada!"
        mostrar_mensaje_exito(mensaje, label_mensaje)
    except Exception as e:
        messagebox.showerror("Error", f"Error al agregar diferencia: {str(e)}")

def main():
    root = tk.Tk()
    root.title("Gestionador de artículos")
    root.geometry("600x400")

    # Variable para guardar el archivo seleccionado
    archivo_excel = tk.StringVar(value="")

    # --- Logo en la esquina superior derecha
    logo = tk.PhotoImage(file="marolio_logo.png").subsample(5, 5)
    lbl_logo = ttk.Label(root, image=logo)
    lbl_logo.image = logo 
    lbl_logo.place(relx=1.0, y=0, anchor="ne", x=-10)  

    # Label para mostrar archivo seleccionado
    lbl_archivo = ttk.Label(root, text="Archivo: (ninguno seleccionado)")

    # --- Botón para seleccionar archivo
    btn_archivo = ttk.Button(
        root,
        text="Buscar archivo Excel",
        command=lambda: seleccionar_archivo(archivo_excel, lbl_archivo)
    )

    btn_archivo.pack(pady=10)
    lbl_archivo.pack()

    # Label para mensajes de estado
    label_mensaje = ttk.Label(root, text="", font=('Arial', 10))
    label_mensaje.pack(side="bottom",pady=30)

    # --- Combobox para elegir el modo
    ttk.Label(root, text="Elegí el modo:").pack(pady=(20, 5))

    modo_var = tk.StringVar()
    combo_modo = ttk.Combobox(root, textvariable=modo_var, state="readonly")
    combo_modo["values"] = ("Buscar por artículo", "Buscar por pasillo", "Diferencias")
    combo_modo.pack()

    # Contenedor para los formularios
    form_container = ttk.Frame(root, padding=10)
    form_container.pack(fill="both", expand=True, pady=15)

    # --- Formulario por Buscar por artículo
    form_buscar_por_articulo = ttk.Frame(form_container, padding=10)
    ttk.Label(form_buscar_por_articulo, text="Artículo:").grid(row=0, column=0, sticky="w", pady=5)
    entry_articulo = ttk.Entry(form_buscar_por_articulo)
    entry_articulo.grid(row=0, column=1, pady=5, sticky="ew")

    ttk.Button(
        form_buscar_por_articulo, 
        text="Obtener PDF", 
        command=lambda: buscar_por_articulo(entry_articulo, archivo_excel, label_mensaje)
    ).grid(row=2, column=4, columnspan=2, pady=10)

    ttk.Button(
        form_buscar_por_articulo, 
        text="Buscar artículo", 
        command=lambda: buscar_descripcion_articulo(entry_articulo, archivo_excel, label_mensaje)
    ).grid(row=2, column=0, columnspan=2, pady=10)

    # --- Formulario por Buscar por pasillo
    form_buscar_por_pasillo = ttk.Frame(form_container, padding=10)
    ttk.Label(form_buscar_por_pasillo, text="Pasillo:").grid(row=0, column=0, sticky="w", pady=5)
    entry_pasillo = ttk.Entry(form_buscar_por_pasillo)
    entry_pasillo.grid(row=0, column=1, pady=5, sticky="ew")

    ttk.Button(
        form_buscar_por_pasillo, 
        text="Obtener PDF", 
        command=lambda: buscar_por_pasillo(entry_pasillo, archivo_excel, label_mensaje)
    ).grid(row=2, column=0, columnspan=2, pady=10)

    # --- Formulario por Diferencias
    form_diferencias = ttk.Frame(form_container, padding=10)

    ttk.Label(form_diferencias, text="Localizador:").grid(row=0, column=0, sticky="w", pady=5)
    entry_diferencias = ttk.Entry(form_diferencias)
    entry_diferencias.grid(row=0, column=1, pady=5, sticky="ew")

    ttk.Label(form_diferencias, text="Estado:").grid(row=0, column=2, sticky="w", pady=5)

    estado_var = ttk.Combobox(form_diferencias, values=["FALTANTE", "SOBRANTE"], state="readonly")
    estado_var.grid(row=0, column=3, pady=5, sticky="ew")
    estado_var.current(0)

    ttk.Button(
        form_diferencias, 
        text="Agregar diferencia", 
        command=lambda: agregar_diferencias(entry_diferencias, estado_var, archivo_excel, label_mensaje)
    ).grid(row=2, column=0, columnspan=2, pady=10)

    combo_modo.bind(
        "<<ComboboxSelected>>",
        lambda event: mostrar_formulario(form_container, modo_var, form_buscar_por_articulo, form_buscar_por_pasillo, form_diferencias)
    )

    root.mainloop()

if __name__ == "__main__":
    main()